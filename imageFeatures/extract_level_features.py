"""
收纳关卡特征提取工具
自动分析关卡示意图，提取关键特征并保存到Excel
"""

import os
import pandas as pd
import numpy as np
from PIL import Image
import cv2
from pathlib import Path
import json
from typing import Dict, List, Tuple
import warnings
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')


class LevelFeatureExtractor:
    """关卡特征提取器"""
    
    def __init__(self, image_dir: str, output_file: str = "level_features.xlsx"):
        """
        初始化特征提取器
        
        Args:
            image_dir: 图片目录路径
            output_file: 输出Excel文件名
        """
        self.image_dir = Path(image_dir)
        self.output_file = output_file
        self.supported_formats = {'.jpg', '.jpeg', '.png', '.bmp', '.gif'}
        
    def extract_all_features(self, image_path: str) -> Dict:
        """
        提取图像的所有特征
        
        Args:
            image_path: 图片路径
            
        Returns:
            包含所有特征的字典
        """
        try:
            # 读取图像
            img = cv2.imread(str(image_path))
            if img is None:
                raise ValueError(f"无法读取图片: {image_path}")
            
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            h, w = img.shape[:2]
            
            features = {}
            
            # 1. 基础信息
            #features['图片宽度'] = w
            #features['图片高度'] = h
            #features['宽高比'] = round(w / h, 3) if h > 0 else 0
            #features['总像素数'] = w * h
            
            # 2. 颜色特征
            color_features = self._extract_color_features(img_rgb)
            features.update(color_features)
            
            # 3. 纹理特征
            texture_features = self._extract_texture_features(img_gray)
            features.update(texture_features)
            
            # 4. 布局特征
            layout_features = self._extract_layout_features(img_gray, img_rgb)
            features.update(layout_features)
            
            # 5. 复杂度特征
            complexity_features = self._extract_complexity_features(img_gray, img_rgb)
            features.update(complexity_features)
            
            # 6. 边缘特征
            edge_features = self._extract_edge_features(img_gray)
            features.update(edge_features)
            
            return features
            
        except Exception as e:
            print(f"提取特征时出错 {image_path}: {e}")
            return {}
    
    def _extract_color_features(self, img_rgb: np.ndarray) -> Dict:
        """提取颜色特征"""
        features = {}
        
        # 平均RGB值
        features['平均R值'] = int(np.mean(img_rgb[:, :, 0]))
        features['平均G值'] = int(np.mean(img_rgb[:, :, 1]))
        features['平均B值'] = int(np.mean(img_rgb[:, :, 2]))
        
        # 颜色标准差（颜色分布均匀度）
        features['R标准差'] = round(np.std(img_rgb[:, :, 0]), 2)
        features['G标准差'] = round(np.std(img_rgb[:, :, 1]), 2)
        features['B标准差'] = round(np.std(img_rgb[:, :, 2]), 2)
        
        # 亮度
        gray = cv2.cvtColor(img_rgb, cv2.COLOR_RGB2GRAY)
        features['平均亮度'] = round(np.mean(gray), 2)
        features['亮度标准差'] = round(np.std(gray), 2)
        
        # 颜色种类（唯一颜色数量）
        img_reshaped = img_rgb.reshape(-1, 3)
        unique_colors = len(np.unique(img_reshaped, axis=0))
        features['唯一颜色数'] = unique_colors
        features['颜色丰富度'] = round(unique_colors / (img_rgb.shape[0] * img_rgb.shape[1]) * 100, 2)
        
        # 主色调（通过K-means聚类找到主要颜色）
        try:
            from sklearn.cluster import KMeans
            from sklearn.metrics import silhouette_score
            
            pixels = img_reshaped.astype(np.float32)
            
            # 固定3个主色调（保持向后兼容）
            kmeans_3 = KMeans(n_clusters=3, random_state=42, n_init=10)
            kmeans_3.fit(pixels)
            dominant_colors = kmeans_3.cluster_centers_.astype(int)
            features['主色调1_R'] = int(dominant_colors[0][0])
            features['主色调1_G'] = int(dominant_colors[0][1])
            features['主色调1_B'] = int(dominant_colors[0][2])
            
            # 动态确定最佳颜色聚类数量（使用肘部法则或轮廓系数）
            # 尝试2-10个聚类，选择轮廓系数最高的
            best_k = 3
            best_score = -1
            
            # 为了性能，只对采样后的像素进行聚类
            sample_size = min(10000, len(pixels))
            if len(pixels) > sample_size:
                sample_indices = np.random.choice(len(pixels), sample_size, replace=False)
                pixels_sample = pixels[sample_indices]
            else:
                pixels_sample = pixels
            
            for k in range(2, min(11, len(np.unique(pixels_sample, axis=0)) + 1)):
                try:
                    kmeans = KMeans(n_clusters=k, random_state=42, n_init=5)
                    labels = kmeans.fit_predict(pixels_sample)
                    if len(np.unique(labels)) > 1:
                        score = silhouette_score(pixels_sample, labels)
                        if score > best_score:
                            best_score = score
                            best_k = k
                except:
                    continue
            
            features['颜色聚类数量'] = best_k
            features['颜色聚类轮廓系数'] = round(best_score, 3) if best_score > -1 else 0
            
        except:
            features['主色调1_R'] = 0
            features['主色调1_G'] = 0
            features['主色调1_B'] = 0
            features['颜色聚类数量'] = 0
            features['颜色聚类轮廓系数'] = 0
        
        return features
    
    def _extract_texture_features(self, img_gray: np.ndarray) -> Dict:
        """提取纹理特征"""
        features = {}
        
        # 使用LBP (Local Binary Pattern) 计算纹理
        try:
            from skimage.feature import local_binary_pattern
            lbp = local_binary_pattern(img_gray, 8, 1, method='uniform')
            features['纹理均匀度'] = round(np.std(lbp), 2)
        except:
            features['纹理均匀度'] = 0
        
        # 梯度特征（纹理复杂度）
        grad_x = cv2.Sobel(img_gray, cv2.CV_64F, 1, 0, ksize=3)
        grad_y = cv2.Sobel(img_gray, cv2.CV_64F, 0, 1, ksize=3)
        gradient_magnitude = np.sqrt(grad_x**2 + grad_y**2)
        features['平均梯度强度'] = round(np.mean(gradient_magnitude), 2)
        features['梯度标准差'] = round(np.std(gradient_magnitude), 2)
        
        return features
    
    def _extract_layout_features(self, img_gray: np.ndarray, img_rgb: np.ndarray) -> Dict:
        """提取布局特征"""
        features = {}
        h, w = img_gray.shape
        
        # 中心区域亮度 vs 边缘区域亮度
        center_h, center_w = h // 2, w // 2
        margin = min(h, w) // 4
        
        center_region = img_gray[center_h-margin:center_h+margin, 
                                 center_w-margin:center_w+margin]
        edge_region = np.concatenate([
            img_gray[:margin, :].flatten(),
            img_gray[-margin:, :].flatten(),
            img_gray[:, :margin].flatten(),
            img_gray[:, -margin:].flatten()
        ])
        
        if len(center_region) > 0 and len(edge_region) > 0:
            features['中心区域平均亮度'] = round(np.mean(center_region), 2)
            features['边缘区域平均亮度'] = round(np.mean(edge_region), 2)
            features['中心边缘亮度差'] = round(abs(np.mean(center_region) - np.mean(edge_region)), 2)
        else:
            features['中心区域平均亮度'] = 0
            features['边缘区域平均亮度'] = 0
            features['中心边缘亮度差'] = 0
        
        # 对称性（水平对称）
        left_half = img_gray[:, :w//2]
        right_half = cv2.flip(img_gray[:, w//2:], 1)
        if right_half.shape[1] > left_half.shape[1]:
            right_half = right_half[:, :left_half.shape[1]]
        elif right_half.shape[1] < left_half.shape[1]:
            left_half = left_half[:, :right_half.shape[1]]
        
        if left_half.shape == right_half.shape:
            symmetry = 1 - np.mean(np.abs(left_half.astype(float) - right_half.astype(float))) / 255
            features['水平对称度'] = round(symmetry, 3)
        else:
            features['水平对称度'] = 0
        
        # 对称性（垂直对称）
        top_half = img_gray[:h//2, :]
        bottom_half = cv2.flip(img_gray[h//2:, :], 0)
        if bottom_half.shape[0] > top_half.shape[0]:
            bottom_half = bottom_half[:top_half.shape[0], :]
        elif bottom_half.shape[0] < top_half.shape[0]:
            top_half = top_half[:bottom_half.shape[0], :]
        
        if top_half.shape == bottom_half.shape:
            symmetry = 1 - np.mean(np.abs(top_half.astype(float) - bottom_half.astype(float))) / 255
            features['垂直对称度'] = round(symmetry, 3)
        else:
            features['垂直对称度'] = 0
        
        # 四象限特征
        mid_h, mid_w = h // 2, w // 2
        quadrant_1 = img_gray[:mid_h, :mid_w]  # 左上
        quadrant_2 = img_gray[:mid_h, mid_w:]  # 右上
        quadrant_3 = img_gray[mid_h:, :mid_w]  # 左下
        quadrant_4 = img_gray[mid_h:, mid_w:]  # 右下
        
        features['象限1平均亮度'] = round(np.mean(quadrant_1), 2)
        features['象限2平均亮度'] = round(np.mean(quadrant_2), 2)
        features['象限3平均亮度'] = round(np.mean(quadrant_3), 2)
        features['象限4平均亮度'] = round(np.mean(quadrant_4), 2)
        
        # 四象限亮度标准差（衡量四个象限的亮度差异）
        quadrant_means = [
            np.mean(quadrant_1),
            np.mean(quadrant_2),
            np.mean(quadrant_3),
            np.mean(quadrant_4)
        ]
        features['四象限亮度标准差'] = round(np.std(quadrant_means), 2)
        
        # 连通域在四象限的分布（如果有连通域数据）
        try:
            _, binary = cv2.threshold(img_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            num_labels, labels = cv2.connectedComponents(binary)
            if num_labels > 1:
                quadrant_counts = [0, 0, 0, 0]
                for label_id in range(1, num_labels):
                    mask = (labels == label_id).astype(np.uint8)
                    moments = cv2.moments(mask)
                    if moments['m00'] > 0:
                        cx = moments['m10'] / moments['m00']
                        cy = moments['m01'] / moments['m00']
                        # 判断重心在哪个象限
                        if cy < mid_h and cx < mid_w:
                            quadrant_counts[0] += 1
                        elif cy < mid_h and cx >= mid_w:
                            quadrant_counts[1] += 1
                        elif cy >= mid_h and cx < mid_w:
                            quadrant_counts[2] += 1
                        else:
                            quadrant_counts[3] += 1
                
                total_objects = sum(quadrant_counts)
                if total_objects > 0:
                    features['象限1物体占比'] = round(quadrant_counts[0] / total_objects * 100, 2)
                    features['象限2物体占比'] = round(quadrant_counts[1] / total_objects * 100, 2)
                    features['象限3物体占比'] = round(quadrant_counts[2] / total_objects * 100, 2)
                    features['象限4物体占比'] = round(quadrant_counts[3] / total_objects * 100, 2)
                else:
                    features['象限1物体占比'] = 0
                    features['象限2物体占比'] = 0
                    features['象限3物体占比'] = 0
                    features['象限4物体占比'] = 0
            else:
                features['象限1物体占比'] = 0
                features['象限2物体占比'] = 0
                features['象限3物体占比'] = 0
                features['象限4物体占比'] = 0
        except:
            features['象限1物体占比'] = 0
            features['象限2物体占比'] = 0
            features['象限3物体占比'] = 0
            features['象限4物体占比'] = 0
        
        return features
    
    def _extract_complexity_features(self, img_gray: np.ndarray, img_rgb: np.ndarray) -> Dict:
        """提取复杂度特征"""
        features = {}
        
        # 信息熵（复杂度指标）
        hist, _ = np.histogram(img_gray.flatten(), bins=256, range=(0, 256))
        hist = hist[hist > 0]  # 去除0值
        prob = hist / hist.sum()
        entropy = -np.sum(prob * np.log2(prob))
        features['信息熵'] = round(entropy, 3)
        
        # 对比度
        features['对比度'] = round(np.std(img_gray), 2)
        
        # 物体数量（通过连通域分析）
        _, binary = cv2.threshold(img_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        num_labels, labels = cv2.connectedComponents(binary)
        features['连通域数量'] = num_labels - 1  # 减去背景
        
        # 连通域形状特征
        if num_labels > 1:  # 有物体
            h, w = img_gray.shape
            total_pixels = h * w
            areas = []
            aspect_ratios = []
            compactness_scores = []
            centroids = []
            
            for label_id in range(1, num_labels):  # 跳过背景（label 0）
                # 获取该连通域的掩码
                mask = (labels == label_id).astype(np.uint8)
                area = np.sum(mask)
                
                if area > 0:
                    areas.append(area)
                    
                    # 计算边界框
                    coords = np.column_stack(np.where(mask > 0))
                    if len(coords) > 0:
                        y_min, x_min = coords.min(axis=0)
                        y_max, x_max = coords.max(axis=0)
                        bbox_h = y_max - y_min + 1
                        bbox_w = x_max - x_min + 1
                        
                        # 长宽比
                        if bbox_h > 0:
                            aspect_ratio = bbox_w / bbox_h
                            aspect_ratios.append(aspect_ratio)
                        
                        # 圆形度（Compactness）：4π*面积/周长²
                        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                        if len(contours) > 0:
                            perimeter = cv2.arcLength(contours[0], True)
                            if perimeter > 0:
                                compactness = (4 * np.pi * area) / (perimeter ** 2)
                                compactness_scores.append(compactness)
                        
                        # 重心
                        moments = cv2.moments(mask)
                        if moments['m00'] > 0:
                            cx = moments['m10'] / moments['m00']
                            cy = moments['m01'] / moments['m00']
                            centroids.append([cx, cy])
            
            # 连通域面积特征
            if areas:
                features['连通域平均面积'] = round(np.mean(areas), 2)
                features['连通域面积标准差'] = round(np.std(areas), 2)
                features['最大连通域面积占比'] = round(max(areas) / total_pixels * 100, 2)
            else:
                features['连通域平均面积'] = 0
                features['连通域面积标准差'] = 0
                features['最大连通域面积占比'] = 0
            
            # 连通域长宽比特征
            if aspect_ratios:
                features['连通域平均长宽比'] = round(np.mean(aspect_ratios), 3)
                features['连通域长宽比标准差'] = round(np.std(aspect_ratios), 3)
            else:
                features['连通域平均长宽比'] = 0
                features['连通域长宽比标准差'] = 0
            
            # 连通域圆形度特征
            if compactness_scores:
                features['连通域平均圆形度'] = round(np.mean(compactness_scores), 3)
            else:
                features['连通域平均圆形度'] = 0
            
            # 连通域空间分布特征
            if len(centroids) > 1:
                centroids = np.array(centroids)
                # 重心位置（归一化到0-1）
                features['连通域重心X均值'] = round(np.mean(centroids[:, 0]) / w, 3)
                features['连通域重心Y均值'] = round(np.mean(centroids[:, 1]) / h, 3)
                # 重心分布标准差（分散程度）
                features['连通域重心X标准差'] = round(np.std(centroids[:, 0]) / w, 3)
                features['连通域重心Y标准差'] = round(np.std(centroids[:, 1]) / h, 3)
                
                # 计算连通域之间的平均距离
                from scipy.spatial.distance import pdist
                distances = pdist(centroids)
                if len(distances) > 0:
                    features['连通域平均间距'] = round(np.mean(distances), 2)
                else:
                    features['连通域平均间距'] = 0
            elif len(centroids) == 1:
                features['连通域重心X均值'] = round(centroids[0][0] / w, 3)
                features['连通域重心Y均值'] = round(centroids[0][1] / h, 3)
                features['连通域重心X标准差'] = 0
                features['连通域重心Y标准差'] = 0
                features['连通域平均间距'] = 0
            else:
                features['连通域重心X均值'] = 0
                features['连通域重心Y均值'] = 0
                features['连通域重心X标准差'] = 0
                features['连通域重心Y标准差'] = 0
                features['连通域平均间距'] = 0
        else:
            # 没有连通域
            features['连通域平均面积'] = 0
            features['连通域面积标准差'] = 0
            features['最大连通域面积占比'] = 0
            features['连通域平均长宽比'] = 0
            features['连通域长宽比标准差'] = 0
            features['连通域平均圆形度'] = 0
            features['连通域重心X均值'] = 0
            features['连通域重心Y均值'] = 0
            features['连通域重心X标准差'] = 0
            features['连通域重心Y标准差'] = 0
            features['连通域平均间距'] = 0
        
        # 空白区域比例
        white_pixels = np.sum(img_gray > 200)
        total_pixels = img_gray.size
        features['空白区域比例'] = round(white_pixels / total_pixels * 100, 2)
        
        return features
    
    def _extract_edge_features(self, img_gray: np.ndarray) -> Dict:
        """提取边缘特征"""
        features = {}
        
        # Canny边缘检测
        edges = cv2.Canny(img_gray, 50, 150)
        edge_pixels = np.sum(edges > 0)
        total_pixels = edges.size
        features['边缘像素比例'] = round(edge_pixels / total_pixels * 100, 3)
        
        # 边缘密度
        features['边缘密度'] = round(edge_pixels / (img_gray.shape[0] * img_gray.shape[1]), 4)
        
        return features
    
    def process_all_images(self, traffic_data_file: str = None) -> pd.DataFrame:
        """
        处理所有图片并提取特征
        
        Args:
            traffic_data_file: 流量数据文件路径（CSV或Excel），应包含'关卡名称'和'流量好坏'列
            
        Returns:
            包含所有特征的DataFrame
        """
        if not self.image_dir.exists():
            raise ValueError(f"图片目录不存在: {self.image_dir}")
        
        # 读取流量数据（如果提供）
        traffic_data = {}
        if traffic_data_file and os.path.exists(traffic_data_file):
            try:
                if traffic_data_file.endswith('.csv'):
                    df_traffic = pd.read_csv(traffic_data_file)
                else:
                    df_traffic = pd.read_excel(traffic_data_file)
                
                # 尝试不同的列名
                name_col = None
                traffic_col = None
                for col in df_traffic.columns:
                    if '名称' in str(col) or 'name' in str(col).lower() or '关卡' in str(col):
                        name_col = col
                    if '流量' in str(col) or 'traffic' in str(col).lower() or '好坏' in str(col):
                        traffic_col = col
                
                if name_col and traffic_col:
                    for _, row in df_traffic.iterrows():
                        level_name = str(row[name_col])
                        traffic_value = row[traffic_col]
                        traffic_data[level_name] = traffic_value
                    print(f"成功加载 {len(traffic_data)} 条流量数据")
                else:
                    print("警告: 未找到合适的列名，请确保CSV/Excel包含'关卡名称'和'流量好坏'列")
            except Exception as e:
                print(f"读取流量数据时出错: {e}")
        
        # 获取所有图片文件
        image_files = []
        for ext in self.supported_formats:
            image_files.extend(list(self.image_dir.glob(f"*{ext}")))
            image_files.extend(list(self.image_dir.glob(f"*{ext.upper()}")))
        
        # 去重（Windows文件系统不区分大小写，可能导致重复）
        image_files = list(set(image_files))
        
        if not image_files:
            raise ValueError(f"在 {self.image_dir} 中未找到图片文件")
        
        print(f"找到 {len(image_files)} 张图片，开始提取特征...")
        
        # 提取所有图片的特征
        all_features = []
        for img_path in sorted(image_files):
            level_name = img_path.stem  # 文件名（不含扩展名）
            print(f"处理: {level_name}")
            
            features = self.extract_all_features(img_path)
            if features:
                features['关卡名称'] = level_name
                features['图片路径'] = str(img_path)
                
                # 添加流量数据
                if level_name in traffic_data:
                    features['流量好坏'] = traffic_data[level_name]
                else:
                    features['流量好坏'] = None
                
                all_features.append(features)
        
        # 转换为DataFrame
        df = pd.DataFrame(all_features)
        
        # 重新排列列的顺序，将关键列放在前面
        if not df.empty:
            priority_cols = ['关卡名称', '流量好坏', '图片路径']
            other_cols = [col for col in df.columns if col not in priority_cols]
            df = df[priority_cols + other_cols]
        
        return df
    
    def save_to_excel(self, df: pd.DataFrame, sheet_name: str = "关卡特征"):
        """保存到Excel文件"""
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 自动调整列宽
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, start=1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                # 使用 get_column_letter 正确处理超过26列的情况
                column_letter = get_column_letter(idx)
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        print(f"\n特征已保存到: {self.output_file}")
        print(f"共 {len(df)} 个关卡，{len(df.columns)} 个特征")


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='收纳关卡特征提取工具')
    parser.add_argument('--image_dir', type=str, default='images', 
                       help='图片目录路径 (默认: images)')
    parser.add_argument('--traffic_file', type=str, default=None,
                       help='流量数据文件路径 (CSV或Excel，可选)')
    parser.add_argument('--output', type=str, default='level_features.xlsx',
                       help='输出Excel文件名 (默认: level_features.xlsx)')
    
    args = parser.parse_args()
    
    # 创建提取器
    extractor = LevelFeatureExtractor(args.image_dir, args.output)
    
    # 处理所有图片
    df = extractor.process_all_images(args.traffic_file)
    
    # 保存到Excel
    extractor.save_to_excel(df)
    
    # 如果有流量数据，显示统计信息
    if '流量好坏' in df.columns and df['流量好坏'].notna().any():
        print("\n=== 流量分析 ===")
        print(df['流量好坏'].value_counts())
        
        # 对比好和坏关卡的特征差异
        good_levels = df[df['流量好坏'] == '好'] if '好' in df['流量好坏'].values else pd.DataFrame()
        bad_levels = df[df['流量好坏'] == '坏'] if '坏' in df['流量好坏'].values else pd.DataFrame()
        
        if not good_levels.empty and not bad_levels.empty:
            print("\n=== 好关卡 vs 坏关卡特征对比 ===")
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            for col in numeric_cols[:10]:  # 显示前10个数值特征
                if col not in ['图片宽度', '图片高度', '总像素数']:
                    good_mean = good_levels[col].mean()
                    bad_mean = bad_levels[col].mean()
                    diff = good_mean - bad_mean
                    print(f"{col}: 好={good_mean:.2f}, 坏={bad_mean:.2f}, 差异={diff:.2f}")


if __name__ == "__main__":
    main()

